from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

@dataclass
class PlaywrightTestResult:
    focus: str
    test_type: str
    test_id: str
    pre_condition: str
    scenario: str
    steps: str
    expected_result: str
    result: str  # PASS, FAIL, SKIPPED
    test_script: str = ""  
    notes: str = ""

class PlaywrightReportService:
    def __init__(self):
        self.columns = [
            "Focus", "Type", "ID", "Pre-Condition", 
            "Scenario", "Test Steps", "Expected Result", 
            "Result", "Test Script", "Notes / Issue"
        ]

    def generate_report(self, results: list[PlaywrightTestResult], output_path: Path) -> Path:
        """Generates a QA/TSD style Excel report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Execution Report"

        # Headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        for col, header in enumerate(self.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data
        for row_idx, result in enumerate(results, 2):
            data = [
                result.focus, result.test_type, result.test_id, 
                result.pre_condition, result.scenario, result.steps, 
                result.expected_result, result.result, result.test_script, result.notes
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Conditional Formatting for Result column
                if col_idx == 8: # Result column
                    if value.upper() == "PASS":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
                        cell.font = Font(color="006100")
                    elif value.upper() == "FAIL":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
                        cell.font = Font(color="9C0006")
                    elif value.upper() == "SKIPPED":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
                        cell.font = Font(color="9C6500")

        # Auto-size columns and wrap text
        for col in range(1, len(self.columns) + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 20 # Base width
            if col in [5, 6, 7, 9]: # Scenario, Steps, Expected, Notes
                ws.column_dimensions[letter].width = 40

        wb.save(output_path)
        return output_path

    def parse_playwright_json(self, json_data: dict) -> list[PlaywrightTestResult]:
        """
        Optional: Parse Playwright's JSON reporter output to populate results.
        For now, we'll manually create results or mock this.
        """
        results = []
        # Logic to parse Playwright JSON output would go here
        return results
