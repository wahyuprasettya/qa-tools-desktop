from __future__ import annotations

import csv
from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.exceptions import ExportError


class ExportService:
    def export(
        self,
        dataframe: pd.DataFrame,
        output_path: Path,
        file_type: str,
        progress_callback: Callable[[int, str], None] | None = None,
    ) -> Path:
        if dataframe.empty:
            raise ExportError("There is no table data to export.")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if file_type == "xlsx":
            return self._export_xlsx(dataframe, output_path, progress_callback)
        if file_type == "csv":
            return self._export_csv(dataframe, output_path, progress_callback)
        raise ExportError(f"Unsupported export format: {file_type}")

    def _export_csv(
        self,
        dataframe: pd.DataFrame,
        output_path: Path,
        progress_callback: Callable[[int, str], None] | None,
    ) -> Path:
        self._report_progress(progress_callback, 0, "Preparing CSV export...")
        total_rows = len(dataframe)
        with output_path.open("w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(dataframe.columns)
            self._report_progress(progress_callback, 5, "Writing CSV header...")
            for row_number, row in enumerate(dataframe.itertuples(index=False, name=None), start=1):
                writer.writerow(["" if pd.isna(value) else value for value in row])
                percent = 5 + int((row_number / total_rows) * 90)
                self._report_progress(progress_callback, min(percent, 95), f"Writing CSV row {row_number}/{total_rows}...")
        self._report_progress(progress_callback, 100, "CSV export complete.")
        return output_path

    def _export_xlsx(
        self,
        dataframe: pd.DataFrame,
        output_path: Path,
        progress_callback: Callable[[int, str], None] | None,
    ) -> Path:
        self._report_progress(progress_callback, 0, "Preparing XLSX workbook...")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clean Table"
        sheet.freeze_panes = "A2"

        for column_number, column_name in enumerate(dataframe.columns, start=1):
            sheet.cell(row=1, column=column_number, value=str(column_name))
        self._report_progress(progress_callback, 10, "Writing XLSX header...")

        total_rows = len(dataframe)
        for row_number, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
            for column_number, value in enumerate(row, start=1):
                sheet.cell(row=row_number, column=column_number, value="" if pd.isna(value) else value)
            written_rows = row_number - 1
            percent = 10 + int((written_rows / total_rows) * 55)
            self._report_progress(
                progress_callback,
                min(percent, 65),
                f"Writing XLSX row {written_rows}/{total_rows}...",
            )

        header_fill = PatternFill("solid", fgColor="1F6FEB")
        zebra_fill = PatternFill("solid", fgColor="F6F8FA")
        header_font = Font(bold=True, color="FFFFFF")

        self._report_progress(progress_callback, 70, "Styling XLSX header...")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self._report_progress(progress_callback, 78, "Styling XLSX rows...")
        for row in sheet.iter_rows(min_row=2):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = zebra_fill
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)

        self._report_progress(progress_callback, 90, "Adding XLSX filters...")
        sheet.auto_filter.ref = sheet.dimensions
        ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName="CleanTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        self._report_progress(progress_callback, 95, "Saving XLSX file...")
        workbook.save(output_path)
        self._report_progress(progress_callback, 100, "XLSX export complete.")
        return output_path

    def _report_progress(
        self,
        progress_callback: Callable[[int, str], None] | None,
        percent: int,
        message: str,
    ) -> None:
        if progress_callback is not None:
            progress_callback(max(0, min(percent, 100)), message)
