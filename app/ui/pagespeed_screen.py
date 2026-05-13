from __future__ import annotations

from PySide6.QtCore import Qt, QThreadPool
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QProgressBar,
    QPushButton,
    QTableView,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QFormLayout,
    QFileDialog,
    QMessageBox,
)

from app.services.config_manager import AppSettings
from app.models.pagespeed_model import PageSpeedTableModel
from app.services.pagespeed_service import PageSpeedService, PageSpeedSignals
from app.utils.worker import Worker


class PageSpeedScreen(QWidget):
    def __init__(self, settings: AppSettings) -> None:
        super().__init__()
        self.settings = settings
        self.service = PageSpeedService()
        self.thread_pool = QThreadPool.globalInstance()
        self.is_running = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 28, 28, 28)
        layout.setSpacing(16)

        # Header
        title = QLabel("Accessibility Test")
        title.setObjectName("PageTitle")
        subtitle = QLabel("Analyze your web app's accessibility using Google PageSpeed API.")
        subtitle.setObjectName("Muted")
        layout.addWidget(title)
        layout.addWidget(subtitle)

        # Input Form
        form_frame = QFrame()
        form_frame.setObjectName("MetricCard")
        form_layout = QFormLayout(form_frame)
        
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("https://example.com")
        
        form_layout.addRow("Target URL:", self.url_input)
        
        self.run_btn = QPushButton("Run Test")
        self.run_btn.setObjectName("PrimaryButton")
        self.run_btn.clicked.connect(self._run_test)
        
        self.export_btn = QPushButton("Export Report")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export_report)
        
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.run_btn)
        btn_layout.addWidget(self.export_btn)
        
        form_layout.addRow("", btn_layout)
        
        layout.addWidget(form_frame)

        # Dashboard / Score
        self.score_card = QFrame()
        self.score_card.setObjectName("MetricCard")
        score_layout = QVBoxLayout(self.score_card)
        score_title = QLabel("Overall Accessibility Score")
        score_layout.addWidget(score_title, alignment=Qt.AlignCenter)
        self.score_value = QLabel("- / 100")
        self.score_value.setStyleSheet("font-size: 32px; font-weight: bold; color: #9ca3af;")
        score_layout.addWidget(self.score_value, alignment=Qt.AlignCenter)
        layout.addWidget(self.score_card)

        # Content Splitter (Table & Terminal)
        content_layout = QHBoxLayout()
        
        # Left side: Table
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Audit Details:"))
        self.model = PageSpeedTableModel()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch) # Description stretches
        left_layout.addWidget(self.table)
        content_layout.addLayout(left_layout, stretch=2)
        
        # Right side: Terminal
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("Execution Logs:"))
        self.terminal = QTextEdit()
        self.terminal.setReadOnly(True)
        font = QFont("monospace")
        font.setStyleHint(QFont.Monospace)
        self.terminal.setFont(font)
        self.terminal.setStyleSheet("background-color: #1e1e1e; color: #d4d4d4; padding: 8px;")
        right_layout.addWidget(self.terminal)
        content_layout.addLayout(right_layout, stretch=1)
        
        layout.addLayout(content_layout, stretch=1)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        layout.addWidget(self.progress)

    def _run_test(self) -> None:
        url = self.url_input.text().strip()
        if not url:
            self.terminal.append("[ERROR] Please enter a valid URL.")
            return

        if not url.startswith("http"):
            url = f"https://{url}"
            self.url_input.setText(url)

        self.model.clear()
        self.terminal.clear()
        self.progress.setValue(0)
        self.score_value.setText("- / 100")
        self.score_value.setStyleSheet("font-size: 32px; font-weight: bold; color: #9ca3af;")
        
        self.run_btn.setEnabled(False)
        self.is_running = True

        api_key = self.settings.pagespeed_api_key.strip()

        signals = PageSpeedSignals()
        signals.progress.connect(self.progress.setValue)
        signals.log.connect(self.terminal.append)
        signals.finished.connect(self._handle_finished)

        # Start service in background thread
        self.worker = Worker(self.service.test_accessibility, url, api_key, signals)
        self.thread_pool.start(self.worker)

    def _handle_finished(self, success: bool, data: dict, error_message: str) -> None:
        self.is_running = False
        self.run_btn.setEnabled(True)
        self.export_btn.setEnabled(success)
        
        if success:
            self.last_results = data # Store for export
            self.last_url = self.url_input.text().strip()
            score = data.get("score", 0)
            self.score_value.setText(f"{score} / 100")
            
            # Color code score
            if score >= 90:
                color = "#10b981" # Green
            elif score >= 50:
                color = "#f59e0b" # Orange
            else:
                color = "#ef4444" # Red
            self.score_value.setStyleSheet(f"font-size: 32px; font-weight: bold; color: {color};")
            
            self.model.set_audits(data.get("audits", []))
            self.terminal.append("\nTest completed successfully.")
        else:
            self.terminal.append(f"\n[FATAL ERROR] Test failed: {error_message}")
            self.score_value.setText("Error")
            self.score_value.setStyleSheet("font-size: 32px; font-weight: bold; color: #ef4444;")
            
        self.progress.setValue(100)

    def _export_report(self) -> None:
        if not hasattr(self, 'last_results') or not self.last_results:
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Report", f"PageSpeed_Accessibility_{self.last_url.replace('https://', '').replace('/', '_')}.xlsx",
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        
        if not file_path:
            return

        try:
            import pandas as pd
            
            audits = self.last_results.get("audits", [])
            report_data = []
            
            for audit in audits:
                # Determine Result
                if audit.score is None:
                    result = "N/A / Manual"
                elif audit.score >= 0.9:
                    result = "PASSED"
                else:
                    result = "FAILED"

                report_data.append({
                    "Focus": "Accessibility",
                    "Type": "Automated (PageSpeed)",
                    "ID": audit.id,
                    "Pre-Condition": f"URL: {self.last_url}",
                    "Scenario": audit.title,
                    "Test Steps": "1. Open URL in PageSpeed Insights\n2. Analyze Accessibility Category\n3. Capture audit result",
                    "Expected Result": audit.description,
                    "Result": result,
                    "Notes / Issue": audit.display_value if audit.display_value else ("Score: " + str(audit.score) if audit.score is not None else "")
                })
            
            df = pd.DataFrame(report_data)
            
            if file_path.endswith('.xlsx'):
                df.to_excel(file_path, index=False)
                
                # Apply styling
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Font, Alignment
                wb = load_workbook(file_path)
                ws = wb.active
                header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Auto-adjust columns
                from openpyxl.utils import get_column_letter
                for col in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 25
                
                wb.save(file_path)
            else:
                df.to_csv(file_path, index=False)
                
            QMessageBox.information(self, "Success", f"Report exported successfully to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export report: {str(e)}")
